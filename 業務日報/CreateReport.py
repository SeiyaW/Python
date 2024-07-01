import datetime
import openpyxl as ox # 必須ライブラリ

class CreateReport:
    def __init__(self, filename, filetype):
        # 現在時刻の取得
        self.__now = datetime.datetime.now()

        # ファイル名や種類の設定
        # 読み込むファイルと作成するファイル
        self.__filename = filename
        self.__filetype = filetype
        self.__loadfile = self.__filename + self.__filetype
        self.__outfile  = self.__filename + self.__now.strftime('%Y年%m月%d日') + self.__filetype

    # ファイルを指定し、コピーを作成する 
    def create_report_excel(self):
        # wb = openpyxl.load_workbook(self.__loadfile)
        wb = ox.load_workbook(self.__loadfile)
        wb.save(self.__outfile)
        return
    
    # 作成したファイルのシート名変更
    def change_sheet(self):
        wb = ox.load_workbook(self.__outfile)
        ws = wb.worksheets[0]
        ws.title = self.__now.strftime('%d日')
        wb.save(self.__outfile)
        return

CR = CreateReport('./【業務日報】', '.xlsx')
CR.create_report_excel()
CR.change_sheet()